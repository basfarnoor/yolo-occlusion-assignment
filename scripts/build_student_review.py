"""Build results/student_review.xlsx: one row per organized sample-stage image,
ready for the student to manually track one target object through the five
occlusion stages. See STUDENT_YOLO_OCCLUSION_ASSIGNMENT.md Task 5.

Does not read or alter the raw YOLO CSV outputs -- it only reuses the image
list from all_images.csv to know which (sample, stage) rows exist.
"""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

PROJECT_ROOT = Path(__file__).resolve().parent.parent
IMAGES_CSV = PROJECT_ROOT / "results" / "all_images.csv"
OUT_PATH = PROJECT_ROOT / "results" / "student_review.xlsx"

STAGE_DISPLAY_NAMES = {
    "previous_no_occlusion": "Previous No Occlusion",
    "first_partial_occlusion": "First Partial Occlusion",
    "full_occlusion": "Full Occlusion",
    "first_partial_appearance": "First Partial Appearance",
    "full_appearance": "Full Appearance",
}

HEADERS = [
    "Sample number", "Stage number", "Stage name", "Image filename",
    "Target description", "Expected target class", "Target detected: Yes or No",
    "Predicted class", "Correct class: Yes, No, or Not applicable",
    "Target confidence", "Failure type", "Student notes",
]

DETECTED_OPTIONS = ["Yes", "No"]
CORRECT_CLASS_OPTIONS = ["Yes", "No", "Not applicable"]
FAILURE_TYPE_OPTIONS = [
    "None", "Missed while visible", "Missed while partially visible",
    "Wrong class", "Box on the wrong object",
    "False target detection during full occlusion", "Unclear",
]

BAND_COLORS = ["FFFFFF", "EAF1FB"]


def load_rows() -> list[dict]:
    with open(IMAGES_CSV, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def main() -> None:
    rows = load_rows()
    if not rows:
        print(f"No rows found in {IMAGES_CSV}; nothing to build.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Target Review"

    ws.append(HEADERS)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    band_index = 0
    prev_sample = None
    for row in rows:
        sample_number = row["sample_number"]
        if sample_number != prev_sample:
            band_index = 1 - band_index
            prev_sample = sample_number

        stage_name_display = STAGE_DISPLAY_NAMES.get(row["stage_name"], row["stage_name"])
        excel_row = [
            sample_number,
            int(row["stage_number"]),
            stage_name_display,
            row["image_filename"],
            "",  # Target description -- filled by student
            "",  # Expected target class -- filled by student
            "",  # Target detected -- dropdown
            "",  # Predicted class -- filled by student
            "",  # Correct class -- dropdown
            "",  # Target confidence -- filled by student
            "",  # Failure type -- dropdown
            "",  # Student notes
        ]
        ws.append(excel_row)
        r = ws.max_row
        fill = PatternFill("solid", fgColor=BAND_COLORS[band_index])
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row=r, column=col_idx).fill = fill

    last_row = ws.max_row

    # Dropdowns
    def add_dropdown(col_letter: str, options: list[str]):
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(options) + '"',
            allow_blank=True,
            showDropDown=False,
        )
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{last_row}")

    add_dropdown("G", DETECTED_OPTIONS)          # Target detected
    add_dropdown("I", CORRECT_CLASS_OPTIONS)     # Correct class
    add_dropdown("K", FAILURE_TYPE_OPTIONS)      # Failure type

    # Column widths
    widths = [14, 12, 24, 34, 26, 20, 20, 16, 26, 16, 34, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{last_row}"

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH} with {last_row - 1} stage rows across "
          f"{len({r['sample_number'] for r in rows})} samples.")


if __name__ == "__main__":
    main()
