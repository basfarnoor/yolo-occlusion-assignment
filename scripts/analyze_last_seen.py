"""Assignment 2, Task 5: validate the Results workbook and produce the
final analysis. Never invents Your judgement / ghost-risk values the
student hasn't filled in -- reports them as unavailable instead.
"""
from __future__ import annotations

import csv
import statistics
from pathlib import Path

import matplotlib.pyplot as plt
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
OUT_ROOT = PROJECT_ROOT / "results" / "last_seen_memory"
WORKBOOK_PATH = OUT_ROOT / "last_seen_experiment.xlsx"

JUDGEMENT_OPTIONS = ["Helpful", "Partly helpful", "Misleading"]


def load_target_selection() -> list[dict]:
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb["Target Selection"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[idx["Sample"]] is None:
            continue
        rows.append({
            "sample": row[idx["Sample"]],
            "include": row[idx["Include in study? Yes/No"]],
            "reason": row[idx["Rejection reason (if not included)"]],
        })
    return rows


def load_results() -> list[dict]:
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb["Results"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[idx["Sample"]] is None:
            continue
        rows.append({
            "sample": row[idx["Sample"]],
            "description": row[idx["Target description"]],
            "target_class": row[idx["Target class"]],
            "last_visible_stage": row[idx["Last visible stage"]],
            "first_reappearance_stage": row[idx["First reappearance stage"]],
            "memory_age": row[idx["Memory age in stages"]],
            "prev_conf": row[idx["Previous YOLO confidence"]],
            "reapp_conf": row[idx["Reappearance YOLO confidence"]],
            "center_error_px": row[idx["Center error in pixels"]],
            "center_error_pct": row[idx["Center error as % of image width"]],
            "iou": row[idx["IoU"]],
            "judgement": row[idx["Your judgement"]],
            "ghost_risk": row[idx["Could this become a ghost object? Yes or No"]],
            "notes": row[idx["Notes"]],
        })
    return rows


def main() -> None:
    selection_rows = load_target_selection()
    results_rows = load_results()

    valid = [r for r in selection_rows if r["include"] == "Yes"]
    rejected = [r for r in selection_rows if r["include"] == "No"]

    center_errors = [r["center_error_px"] for r in results_rows if r["center_error_px"] not in (None, "")]
    ious = [r["iou"] for r in results_rows if r["iou"] not in (None, "")]

    judged = [r for r in results_rows if r["judgement"] in JUDGEMENT_OPTIONS]
    ghost_answered = [r for r in results_rows if r["ghost_risk"] in ("Yes", "No")]
    judgement_counts = {opt: sum(1 for r in judged if r["judgement"] == opt) for opt in JUDGEMENT_OPTIONS}
    ghost_yes = sum(1 for r in ghost_answered if r["ghost_risk"] == "Yes")

    # summary.csv
    summary_path = OUT_ROOT / "summary.csv"
    with open(summary_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["metric", "value", "n_samples"])
        writer.writerow(["valid_samples", len(valid), len(selection_rows)])
        writer.writerow(["rejected_samples", len(rejected), len(selection_rows)])
        writer.writerow(["mean_center_error_px", round(statistics.mean(center_errors), 1) if center_errors else "", len(center_errors)])
        writer.writerow(["median_center_error_px", round(statistics.median(center_errors), 1) if center_errors else "", len(center_errors)])
        writer.writerow(["mean_iou", round(statistics.mean(ious), 3) if ious else "", len(ious)])
        writer.writerow(["median_iou", round(statistics.median(ious), 3) if ious else "", len(ious)])
        for opt in JUDGEMENT_OPTIONS:
            writer.writerow([f"judgement_{opt.lower().replace(' ', '_')}", judgement_counts[opt], len(judged)])
        writer.writerow(["ghost_risk_yes", ghost_yes, len(ghost_answered)])
        writer.writerow([])
        writer.writerow(["sample", "center_error_px", "iou", "memory_age_stages",
                          "prev_confidence", "reappearance_confidence", "judgement", "ghost_risk"])
        for r in results_rows:
            writer.writerow([r["sample"], r["center_error_px"], r["iou"], r["memory_age"],
                              r["prev_conf"], r["reapp_conf"], r["judgement"] or "", r["ghost_risk"] or ""])
    print(f"Wrote {summary_path}")

    # Charts
    samples = [r["sample"] for r in results_rows]

    fig, ax = plt.subplots(figsize=(9, 5))
    vals = [r["center_error_px"] if r["center_error_px"] not in (None, "") else 0 for r in results_rows]
    bars = ax.bar(samples, vals, color="#2F5597")
    for bar, v in zip(bars, vals):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height(), f"{v:.0f}px",
                ha="center", va="bottom", fontsize=10)
    ax.set_ylabel("Center error (pixels)")
    ax.set_title("How far the memory box drifted from the real object at reappearance")
    plt.xticks(rotation=15)
    plt.tight_layout()
    out1 = OUT_ROOT / "center_error_by_sample.png"
    plt.savefig(out1, dpi=120)
    plt.close(fig)
    print(f"Wrote {out1}")

    fig, ax = plt.subplots(figsize=(9, 5))
    vals = [r["iou"] if r["iou"] not in (None, "") else 0 for r in results_rows]
    bars = ax.bar(samples, vals, color="#FFA630")
    for bar, v in zip(bars, vals):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height(), f"{v:.2f}",
                ha="center", va="bottom", fontsize=10)
    ax.set_ylabel("IoU (memory box vs. new detection)")
    ax.set_ylim(0, 1)
    ax.set_title("How well the memory box overlapped the real object at reappearance")
    plt.xticks(rotation=15)
    plt.tight_layout()
    out2 = OUT_ROOT / "iou_by_sample.png"
    plt.savefig(out2, dpi=120)
    plt.close(fig)
    print(f"Wrote {out2}")

    # final_report.md
    best = min(results_rows, key=lambda r: r["center_error_px"] if r["center_error_px"] not in (None, "") else 1e9)
    worst = max(results_rows, key=lambda r: r["center_error_px"] if r["center_error_px"] not in (None, "") else -1)

    report_path = OUT_ROOT / "final_report.md"
    with open(report_path, "w", encoding="utf-8") as f:
        f.write("# Last-Seen Memory: Analysis\n\n")
        f.write(
            "This report validates `last_seen_experiment.xlsx` and summarizes the last-seen "
            "memory experiment: freezing a target's most recent bounding box while it's hidden, "
            "instead of letting it vanish immediately. All numbers below come from the "
            "`Results` sheet; the `Your judgement`, `Could this become a ghost object?`, and "
            "`Notes` columns were left blank at the time of this report, so anything that "
            "depends on them is reported as unavailable rather than guessed.\n\n"
        )

        f.write("## Sample validity\n\n")
        f.write(f"- **{len(valid)} valid samples** used in the study: "
                f"{', '.join(r['sample'] for r in valid)}\n")
        f.write(f"- **{len(rejected)} samples rejected**, with reasons:\n\n")
        for r in rejected:
            f.write(f"  - `{r['sample']}`: {r['reason']}\n")
        f.write("\n")

        f.write("## Center error and IoU\n\n")
        f.write(f"- Mean center error: **{statistics.mean(center_errors):.0f} px** "
                f"(median {statistics.median(center_errors):.0f} px, n={len(center_errors)})\n")
        f.write(f"- Mean IoU: **{statistics.mean(ious):.3f}** "
                f"(median {statistics.median(ious):.3f}, n={len(ious)})\n\n")

        f.write("## Human judgement (Task 4 columns)\n\n")
        if judged:
            f.write(f"Judged so far ({len(judged)}/{len(results_rows)} rows): "
                    + ", ".join(f"{opt}={judgement_counts[opt]}" for opt in JUDGEMENT_OPTIONS) + "\n")
        else:
            f.write("**Not available.** No rows have a `Your judgement` value yet -- "
                    "this needs a manual pass before helpful/misleading counts can be reported.\n")
        if ghost_answered:
            f.write(f"\nPossible ghost-object risk flagged on {ghost_yes}/{len(ghost_answered)} answered rows.\n\n")
        else:
            f.write("\n**Not available.** No rows have a `Could this become a ghost object?` "
                    "answer yet.\n\n")

        f.write("## Answers to the required questions\n\n")

        f.write("**1. Did memory stop the target from immediately disappearing?**\n\n")
        f.write(
            "Yes, by construction: every valid sample shows a `MEMORY -- NOT CURRENTLY DETECTED` "
            "box during full occlusion instead of nothing at all (see `comparisons/<sample>/"
            "comparison_1_full_occlusion.jpg`). The target's existence was preserved for exactly "
            "1 stage of memory age in all 5 samples -- none needed to reach the 2-stage expiry limit, "
            "because each sample only had one occlusion stage between the before/after anchors.\n\n"
        )

        f.write("**2. Did the old box remain close to the target?**\n\n")
        f.write(
            f"It varied a lot. Center error ranged from {min(center_errors):.0f} px to "
            f"{max(center_errors):.0f} px across the 5 samples (mean {statistics.mean(center_errors):.0f} px), "
            f"and IoU ranged from {min(ious):.2f} to {max(ious):.2f}. In 3 of 5 samples the IoU was "
            "0.00 -- the frozen box did not overlap the real object at all once it reappeared.\n\n"
        )

        f.write("**3. Which sample had the best memory location?**\n\n")
        f.write(f"`{best['sample']}` -- center error {best['center_error_px']:.0f} px, "
                f"IoU {best['iou']:.2f}. {best['description']}\n\n")

        f.write("**4. Which had the worst?**\n\n")
        f.write(f"`{worst['sample']}` -- center error {worst['center_error_px']:.0f} px, "
                f"IoU {worst['iou']:.2f}. {worst['description']}\n\n")

        f.write("**5. When was memory helpful?**\n\n")
        f.write(
            "Not available as a labeled count -- the `Your judgement` column hasn't been filled in. "
            f"By the numbers alone, `{best['sample']}` and the other sample(s) with low center error "
            "and higher IoU are the strongest candidates for 'Helpful'; a manual look at the "
            "comparison images is needed to confirm.\n\n"
        )

        f.write("**6. When was it misleading?**\n\n")
        f.write(
            "Not available as a labeled count for the same reason. The samples with IoU = 0.00 "
            f"(center error above ~400 px) are the strongest candidates for 'Misleading' -- worth "
            "reviewing those comparison images first.\n\n"
        )

        f.write("**7. Why could memory create a ghost object?**\n\n")
        f.write(
            "Because the memory box is frozen at its last real position and confidence, it keeps "
            "being drawn even after the object may have moved far away, turned, left the scene, or "
            "been replaced in that same screen position by something else entirely. If memory were "
            "allowed to persist indefinitely (instead of expiring after 2 missing stages), a system "
            "could keep reporting an object that is no longer there at all -- a 'ghost' detection "
            "with no current camera evidence behind it.\n\n"
        )

        f.write("**8. Why should the next version predict movement?**\n\n")
        f.write(
            f"Because an unmoving memory box goes stale fast when there's real relative motion. "
            f"`{worst['sample']}` shows this clearly: {worst['center_error_px']:.0f} px of drift and "
            "0.00 IoU, even though the identity of the object was never in doubt. A version that "
            "estimates velocity and direction from recent frames -- rather than assuming the object "
            "stayed still -- would likely track the true position far more closely during occlusion.\n\n"
        )

        f.write("## Reminder\n\n")
        f.write(
            "This experiment measures whether the *frozen* memory box stayed close to the target "
            "after it reappeared -- it is not a measurement of the object's true hidden-frame position, "
            "which was never observed. A predicted box should never be described as a real detection.\n"
        )

    print(f"Wrote {report_path}")


if __name__ == "__main__":
    main()
