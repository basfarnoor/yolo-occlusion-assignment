"""Assignment 2, Tasks 2-4: last-seen memory tracking, comparison images,
and the Results worksheet.

Memory rule (intentionally simple, per assignments/2_last_seen_memory.md):
  - When YOLO sees the target: draw it green, save its box as "last seen".
  - When YOLO loses it: redraw the saved box unchanged, dashed orange,
    labeled with its memory age (in stages).
  - Memory expires after 2 consecutive missing stages (no motion prediction,
    no Kalman filter, no tracker, no model training).
"""
from __future__ import annotations

import csv
import math
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image, ImageDraw, ImageFont

PROJECT_ROOT = Path(__file__).resolve().parent.parent
OCCLUDED_ROOT = PROJECT_ROOT / "occluded_samples"
DETECTIONS_CSV = PROJECT_ROOT / "results" / "all_detections.csv"
OUT_ROOT = PROJECT_ROOT / "results" / "last_seen_memory"
WORKBOOK_PATH = OUT_ROOT / "last_seen_experiment.xlsx"
COMPARISONS_ROOT = OUT_ROOT / "comparisons"

DISPLAY_CONF_THRESHOLD = 0.25  # same decluttering threshold used for target selection
MATCH_DIST_PX = 200  # same class + center within this distance counts as "still the target"
MAX_AREA_RATIO = 1.6  # reject matches whose box area differs too much (avoids matching the occluder)
MEMORY_EXPIRES_AFTER = 2  # consecutive missing stages

FONT_BOLD = r"C:\Windows\Fonts\arialbd.ttf"
FONT_REG = r"C:\Windows\Fonts\arial.ttf"

STAGE_NAMES = {
    1: "previous_no_occlusion",
    2: "first_partial_occlusion",
    3: "full_occlusion",
    4: "first_partial_appearance",
    5: "full_appearance",
}
STAGE_NAME_TO_NUM = {v: k for k, v in STAGE_NAMES.items()}

GREEN = (0, 200, 0)
ORANGE = (255, 140, 0)
WHITE = (255, 255, 255)
RED = (220, 30, 30)


def load_detections() -> dict[tuple[str, int], list[dict]]:
    from collections import defaultdict
    by_key = defaultdict(list)
    with open(DETECTIONS_CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            key = (row["sample_number"], int(row["stage_number"]))
            by_key[key].append({
                "detected_class": row["detected_class"],
                "confidence": float(row["confidence"]),
                "x1": float(row["x1"]), "y1": float(row["y1"]),
                "x2": float(row["x2"]), "y2": float(row["y2"]),
            })
    return by_key


def find_sample_stages() -> dict[str, list[int]]:
    stages_by_sample = {}
    for sample_dir in sorted(OCCLUDED_ROOT.iterdir()):
        if not sample_dir.is_dir():
            continue
        stages = []
        for img_path in sorted(sample_dir.iterdir()):
            if img_path.suffix.lower() not in (".jpg", ".jpeg", ".png"):
                continue
            stages.append(int(img_path.name.split("_", 1)[0]))
        if stages:
            stages_by_sample[sample_dir.name] = sorted(stages)
    return stages_by_sample


def stage_image_path(sample: str, stage: int) -> Path:
    return OCCLUDED_ROOT / sample / f"{stage}_{STAGE_NAMES[stage]}.jpg"


def resolve_box_by_number(detections_by_key, sample: str, stage: int, box_number: int) -> dict:
    """Reconstruct the same confidence-sorted numbering used in build_target_selection.py."""
    dets = [d for d in detections_by_key.get((sample, stage), []) if d["confidence"] >= DISPLAY_CONF_THRESHOLD]
    dets.sort(key=lambda d: d["confidence"], reverse=True)
    return dets[box_number - 1]


def center(b: dict) -> tuple[float, float]:
    return ((b["x1"] + b["x2"]) / 2, (b["y1"] + b["y2"]) / 2)


def center_dist(a: dict, b: dict) -> float:
    ax, ay = center(a)
    bx, by = center(b)
    return math.hypot(ax - bx, ay - by)


def iou(a: dict, b: dict) -> float:
    ix1, iy1 = max(a["x1"], b["x1"]), max(a["y1"], b["y1"])
    ix2, iy2 = min(a["x2"], b["x2"]), min(a["y2"], b["y2"])
    iw, ih = max(0.0, ix2 - ix1), max(0.0, iy2 - iy1)
    inter = iw * ih
    area_a = (a["x2"] - a["x1"]) * (a["y2"] - a["y1"])
    area_b = (b["x2"] - b["x1"]) * (b["y2"] - b["y1"])
    union = area_a + area_b - inter
    return inter / union if union > 0 else 0.0


def box_area(b: dict) -> float:
    return max(0.0, b["x2"] - b["x1"]) * max(0.0, b["y2"] - b["y1"])


def find_match(detections_by_key, sample: str, stage: int, ref_box: dict):
    """Search a stage's detections for the same class near ref_box's position.

    Requires both a close center distance AND a similar box area, so a large
    occluder of the same class (e.g. the vehicle doing the occluding) doesn't
    get mistaken for the smaller target just because its center is nearby.
    """
    candidates = detections_by_key.get((sample, stage), [])
    ref_area = box_area(ref_box)
    best, best_dist = None, MATCH_DIST_PX
    for d in candidates:
        if d["detected_class"] != ref_box["detected_class"]:
            continue
        area = box_area(d)
        if ref_area <= 0 or area <= 0:
            continue
        area_ratio = max(area, ref_area) / min(area, ref_area)
        if area_ratio > MAX_AREA_RATIO:
            continue
        dist = center_dist(ref_box, d)
        if dist < best_dist:
            best, best_dist = d, dist
    return best, (best_dist if best is not None else None)


def load_valid_targets() -> list[dict]:
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb["Target Selection"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}
    targets = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[idx["Include in study? Yes/No"]] != "Yes":
            continue
        targets.append({
            "sample": row[idx["Sample"]],
            "description": row[idx["Target description"]],
            "before_stage": STAGE_NAME_TO_NUM[row[idx["Before-occlusion stage"]]],
            "after_stage": STAGE_NAME_TO_NUM[row[idx["After-reappearance stage"]]],
            "before_box_num": int(row[idx["Box number before occlusion"]]),
            "after_box_num": int(row[idx["Box number after reappearance"]]),
        })
    return targets


def track_target(detections_by_key, stages_by_sample, target: dict) -> list[dict]:
    """Walk every available stage in order, deciding detected / memory / expired."""
    sample = target["sample"]
    before_stage, after_stage = target["before_stage"], target["after_stage"]
    # Only the range the assignment actually needs: before-occlusion anchor through
    # after-reappearance anchor. Stages outside that range aren't part of this
    # occlusion event and would only add unnecessary auto-match risk.
    stages = [s for s in stages_by_sample[sample] if before_stage <= s <= after_stage]

    anchor_before = resolve_box_by_number(detections_by_key, sample, before_stage, target["before_box_num"])
    anchor_after = resolve_box_by_number(detections_by_key, sample, after_stage, target["after_box_num"])

    records = []
    current_box = None
    consecutive_misses = 0

    for stage in stages:
        if stage == before_stage:
            box, conf = anchor_before, anchor_before["confidence"]
            status = "detected"
            current_box = anchor_before
            consecutive_misses = 0
        elif stage == after_stage:
            box, conf = anchor_after, anchor_after["confidence"]
            status = "detected"
            current_box = anchor_after
            consecutive_misses = 0
        else:
            ref = current_box if current_box is not None else anchor_before
            match, dist = find_match(detections_by_key, sample, stage, ref)
            if match is not None:
                box, conf = match, match["confidence"]
                status = "detected"
                current_box = match
                consecutive_misses = 0
            else:
                consecutive_misses += 1
                if consecutive_misses > MEMORY_EXPIRES_AFTER:
                    box, conf, status = current_box, None, "expired"
                else:
                    box, conf, status = current_box, None, "memory"

        records.append({
            "stage": stage,
            "stage_name": STAGE_NAMES[stage],
            "status": status,
            "box": box,
            "confidence": conf,
            "memory_age": consecutive_misses if status == "memory" else (0 if status == "detected" else None),
        })
    return records


def draw_box(draw: ImageDraw.ImageDraw, box: dict, color, dashed: bool, big_font, small_font, top_label: str,
             sub_label: str = "", label_above: bool = True):
    x1, y1, x2, y2 = box["x1"], box["y1"], box["x2"], box["y2"]
    if dashed:
        dash_len = 14
        x = x1
        while x < x2:
            draw.line([(x, y1), (min(x + dash_len, x2), y1)], fill=color, width=4)
            draw.line([(x, y2), (min(x + dash_len, x2), y2)], fill=color, width=4)
            x += dash_len * 2
        y = y1
        while y < y2:
            draw.line([(x1, y), (x1, min(y + dash_len, y2))], fill=color, width=4)
            draw.line([(x2, y), (x2, min(y + dash_len, y2))], fill=color, width=4)
            y += dash_len * 2
    else:
        draw.rectangle([x1, y1, x2, y2], outline=color, width=4)

    # Labels can go above the box, or -- when a neighboring box's label would
    # collide with that space -- just inside the box's top edge instead.
    label_y = max(0, y1 - 30) if label_above else y1 + 4
    bbox = draw.textbbox((x1, label_y), top_label, font=big_font)
    draw.rectangle([bbox[0] - 3, bbox[1] - 2, bbox[2] + 3, bbox[3] + 2], fill=color)
    draw.text((x1, label_y), top_label, font=big_font, fill=(0, 0, 0) if color == ORANGE else WHITE)
    if sub_label:
        sub_y = bbox[3] + 4
        sbbox = draw.textbbox((x1, sub_y), sub_label, font=small_font)
        draw.rectangle([sbbox[0] - 2, sbbox[1] - 1, sbbox[2] + 2, sbbox[3] + 1], fill=color)
        draw.text((x1, sub_y), sub_label, font=small_font, fill=(0, 0, 0) if color == ORANGE else WHITE)
    return bbox


def build_comparisons(target: dict, records: list[dict]) -> dict:
    """Comparison 1 (occlusion, side by side) and Comparison 2 (reappearance)."""
    sample = target["sample"]
    big_font = ImageFont.truetype(FONT_BOLD, 22)
    small_font = ImageFont.truetype(FONT_REG, 16)
    caption_font = ImageFont.truetype(FONT_BOLD, 20)

    by_stage = {r["stage"]: r for r in records}
    out_dir = COMPARISONS_ROOT / sample
    out_dir.mkdir(parents=True, exist_ok=True)
    results = {}

    # --- Comparison 1: full occlusion, YOLO-only vs YOLO+memory ---
    occl_record = by_stage.get(3)
    if occl_record is not None:
        occl_path = stage_image_path(sample, 3)
        left = Image.open(occl_path).convert("RGB")
        right = left.copy()
        draw_right = ImageDraw.Draw(right)
        if occl_record["status"] in ("memory", "expired") and occl_record["box"] is not None:
            if occl_record["status"] == "memory":
                draw_box(draw_right, occl_record["box"], ORANGE, dashed=True,
                         big_font=big_font, small_font=small_font,
                         top_label="MEMORY -- NOT CURRENTLY DETECTED",
                         sub_label=f"memory age: {occl_record['memory_age']} stage"
                                   f"{'s' if occl_record['memory_age'] != 1 else ''}")
        elif occl_record["status"] == "detected":
            draw_box(draw_right, occl_record["box"], GREEN, dashed=False,
                      big_font=big_font, small_font=small_font,
                      top_label="CURRENT YOLO DETECTION",
                      sub_label=f"{occl_record['box']['detected_class']} {occl_record['confidence']:.2f}")

        combo_w = left.width * 2 + 20
        combo_h = left.height + 90
        combo = Image.new("RGB", (combo_w, combo_h), (25, 25, 25))
        combo.paste(left, (0, 70))
        combo.paste(right, (left.width + 20, 70))
        d = ImageDraw.Draw(combo)
        d.text((10, 10), "LEFT -- YOLO only (no target box if lost)", font=caption_font, fill=WHITE)
        d.text((left.width + 30, 10), "RIGHT -- YOLO plus memory", font=caption_font, fill=WHITE)
        d.text((10, 40), "The orange box is a prediction from memory, not current camera evidence.",
               font=small_font, fill=(255, 220, 150))
        out_path = out_dir / "comparison_1_full_occlusion.jpg"
        combo.save(out_path, quality=92)
        results["comparison_1"] = out_path

    # --- Comparison 2: reappearance, memory box vs new detection ---
    after_stage = target["after_stage"]
    after_record = by_stage.get(after_stage)
    # the memory box carried INTO the reappearance stage is the box from the stage before it
    stage_list = [r["stage"] for r in records]
    idx_after = stage_list.index(after_stage)
    prior_box = records[idx_after - 1]["box"] if idx_after > 0 else None

    if after_record is not None and prior_box is not None:
        img = Image.open(stage_image_path(sample, after_stage)).convert("RGB")
        draw = ImageDraw.Draw(img)
        new_box = after_record["box"]

        # If the two boxes are horizontally close, their above-box labels would
        # collide -- keep the memory label above and tuck the new-detection
        # label just inside its own box instead.
        x_gap = max(prior_box["x1"], new_box["x1"]) - min(prior_box["x2"], new_box["x2"])
        labels_would_collide = x_gap < 220

        draw_box(draw, prior_box, ORANGE, dashed=True, big_font=big_font, small_font=small_font,
                  top_label="OLD MEMORY BOX", sub_label="from before occlusion")
        draw_box(draw, new_box, GREEN, dashed=False, big_font=big_font, small_font=small_font,
                  top_label="NEW YOLO DETECTION",
                  sub_label=f"{new_box['detected_class']} {after_record['confidence']:.2f}",
                  label_above=not labels_would_collide)

        c_old, c_new = center(prior_box), center(new_box)
        draw.line([c_old, c_new], fill=(255, 255, 0), width=3)
        for c in (c_old, c_new):
            draw.ellipse([c[0] - 5, c[1] - 5, c[0] + 5, c[1] + 5], fill=(255, 255, 0))

        center_err = math.hypot(c_old[0] - c_new[0], c_old[1] - c_new[1])
        iou_val = iou(prior_box, new_box)

        caption = (f"Center error: {center_err:.0f} px   |   IoU: {iou_val:.2f}   "
                   f"(IoU near 1 = boxes overlap well; IoU near 0 = remembered location went stale)")
        cap_bbox = draw.textbbox((10, img.height - 40), caption, font=caption_font)
        draw.rectangle([0, cap_bbox[1] - 8, img.width, img.height], fill=(20, 20, 20))
        draw.text((10, img.height - 34), caption, font=caption_font, fill=WHITE)

        out_path = out_dir / "comparison_2_reappearance.jpg"
        img.save(out_path, quality=92)
        results["comparison_2"] = out_path
        results["center_error_px"] = center_err
        results["iou"] = iou_val

    return results


IMAGE_WIDTH_PX = 1600  # confirmed consistent across all CAM_FRONT frames in this dataset

RESULTS_HEADERS = [
    "Sample", "Target description", "Target class",
    "Last visible stage", "First reappearance stage", "Memory age in stages",
    "Previous YOLO confidence", "Reappearance YOLO confidence",
    "Center error in pixels", "Center error as % of image width", "IoU",
    "Your judgement", "Could this become a ghost object? Yes or No", "Notes",
]
JUDGEMENT_OPTIONS = ["Helpful", "Partly helpful", "Misleading"]


def build_results_sheet(all_results: list[dict]) -> None:
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    if "Results" in wb.sheetnames:
        del wb["Results"]
    ws = wb.create_sheet("Results")

    ws.append(RESULTS_HEADERS)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")
    for col_idx in range(1, len(RESULTS_HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    for entry in all_results:
        target, records, comp = entry["target"], entry["records"], entry["comparisons"]
        by_stage = {r["stage"]: r for r in records}
        before_r = by_stage[target["before_stage"]]
        after_r = by_stage[target["after_stage"]]
        # memory age reached just before reappearance = misses counted at the stage right before "after"
        stage_list = [r["stage"] for r in records]
        idx_after = stage_list.index(target["after_stage"])
        memory_age = records[idx_after - 1]["memory_age"] if idx_after > 0 else 0

        center_err = comp.get("center_error_px")
        iou_val = comp.get("iou")

        ws.append([
            target["sample"],
            target["description"],
            before_r["box"]["detected_class"],
            before_r["stage_name"],
            after_r["stage_name"],
            memory_age,
            round(before_r["confidence"], 3),
            round(after_r["confidence"], 3),
            round(center_err, 1) if center_err is not None else "",
            round(center_err / IMAGE_WIDTH_PX * 100, 1) if center_err is not None else "",
            round(iou_val, 3) if iou_val is not None else "",
            "", "", "",
        ])

    last_row = ws.max_row
    dv = DataValidation(type="list", formula1='"' + ",".join(JUDGEMENT_OPTIONS) + '"',
                          allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"L2:L{last_row}")
    dv2 = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv2)
    dv2.add(f"M2:M{last_row}")

    widths = [12, 34, 12, 20, 20, 16, 18, 20, 16, 20, 10, 16, 22, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(RESULTS_HEADERS))}{last_row}"

    wb.save(WORKBOOK_PATH)
    print(f"Wrote 'Results' sheet with {last_row - 1} rows to {WORKBOOK_PATH}")


def main():
    detections_by_key = load_detections()
    stages_by_sample = find_sample_stages()
    targets = load_valid_targets()

    print(f"Tracking {len(targets)} valid targets across their full stage sequences.\n")

    all_results = []
    for target in targets:
        sample = target["sample"]
        records = track_target(detections_by_key, stages_by_sample, target)
        print(f"=== {sample}: {target['description']} ===")
        for r in records:
            conf_str = f"conf={r['confidence']:.2f}" if r["confidence"] is not None else "conf=n/a"
            print(f"  stage {r['stage']} ({r['stage_name']:24s}): {r['status']:8s} {conf_str} "
                  f"memory_age={r['memory_age']}")
        comp = build_comparisons(target, records)
        all_results.append({"target": target, "records": records, "comparisons": comp})
        print()

    build_results_sheet(all_results)
    return all_results


if __name__ == "__main__":
    main()
