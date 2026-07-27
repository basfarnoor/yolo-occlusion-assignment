"""Assignment 2, Task 1: render numbered before/after occlusion images and
build the Target Selection workbook. Never picks the target itself -- that
is the student's job. See assignments/2_last_seen_memory.md.
"""
from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image, ImageDraw, ImageFont

PROJECT_ROOT = Path(__file__).resolve().parent.parent
OCCLUDED_ROOT = PROJECT_ROOT / "occluded_samples"
DETECTIONS_CSV = PROJECT_ROOT / "results" / "all_detections.csv"
OUT_ROOT = PROJECT_ROOT / "results" / "last_seen_memory"
SELECTION_IMAGES_ROOT = OUT_ROOT / "selection_images"
WORKBOOK_PATH = OUT_ROOT / "last_seen_experiment.xlsx"

DISPLAY_CONF_THRESHOLD = 0.25  # decluttering threshold for these selection images only
FONT_PATH = r"C:\Windows\Fonts\arialbd.ttf"

STAGE_NAMES = {
    1: "previous_no_occlusion",
    2: "first_partial_occlusion",
    3: "full_occlusion",
    4: "first_partial_appearance",
    5: "full_appearance",
}

KNOWN_ISSUE_NOTES = {
    "sample_011": "First review left stages 4-5 incomplete (blank/placeholder confidence). Inspect carefully.",
    "sample_012": "First review flagged a possible false detection during full occlusion. Inspect carefully.",
}


def load_detections() -> dict[tuple[str, int], list[dict]]:
    by_key: dict[tuple[str, int], list[dict]] = defaultdict(list)
    with open(DETECTIONS_CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            key = (row["sample_number"], int(row["stage_number"]))
            by_key[key].append({
                "detected_class": row["detected_class"],
                "confidence": float(row["confidence"]),
                "x1": float(row["x1"]), "y1": float(row["y1"]),
                "x2": float(row["x2"]), "y2": float(row["y2"]),
            })
    return by_key


def find_sample_stages() -> dict[str, list[int]]:
    stages_by_sample = {}
    for sample_dir in sorted(OCCLUDED_ROOT.iterdir()):
        if not sample_dir.is_dir():
            continue
        stages = []
        for img_path in sorted(sample_dir.iterdir()):
            if img_path.suffix.lower() not in (".jpg", ".jpeg", ".png"):
                continue
            stage_num = int(img_path.name.split("_", 1)[0])
            stages.append(stage_num)
        if stages:
            stages_by_sample[sample_dir.name] = sorted(stages)
    return stages_by_sample


def stage_image_path(sample: str, stage: int) -> Path:
    return OCCLUDED_ROOT / sample / f"{stage}_{STAGE_NAMES[stage]}.jpg"


def draw_numbered_boxes(image_path: Path, detections: list[dict], out_path: Path) -> list[dict]:
    """Draw boxes with big numbers, sorted by confidence descending.
    Returns the ordered list (index 0 -> box number 1) actually drawn."""
    img = Image.open(image_path).convert("RGB")
    draw = ImageDraw.Draw(img)
    shown = [d for d in detections if d["confidence"] >= DISPLAY_CONF_THRESHOLD]
    shown.sort(key=lambda d: d["confidence"], reverse=True)

    num_font = ImageFont.truetype(FONT_PATH, 34)
    label_font = ImageFont.truetype(FONT_PATH, 16)

    for i, det in enumerate(shown, start=1):
        x1, y1, x2, y2 = det["x1"], det["y1"], det["x2"], det["y2"]
        draw.rectangle([x1, y1, x2, y2], outline=(0, 200, 0), width=3)

        badge_r = 20
        badge_cx, badge_cy = x1 + badge_r + 2, y1 + badge_r + 2
        draw.ellipse([badge_cx - badge_r, badge_cy - badge_r, badge_cx + badge_r, badge_cy + badge_r],
                     fill=(220, 30, 30), outline=(255, 255, 255), width=2)
        num_text = str(i)
        bbox = draw.textbbox((0, 0), num_text, font=num_font)
        tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
        draw.text((badge_cx - tw / 2 - bbox[0], badge_cy - th / 2 - bbox[1]), num_text,
                   font=num_font, fill=(255, 255, 255))

        label = f"{det['detected_class']} {det['confidence']:.2f}"
        label_y = y2 + 2 if y2 + 20 < img.height else max(0, y1 - 20)
        lbox = draw.textbbox((x1, label_y), label, font=label_font)
        draw.rectangle([lbox[0] - 2, lbox[1] - 1, lbox[2] + 2, lbox[3] + 1], fill=(0, 200, 0))
        draw.text((x1, label_y), label, font=label_font, fill=(0, 0, 0))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    img.save(out_path, quality=92)
    return shown


def main() -> None:
    detections_by_key = load_detections()
    stages_by_sample = find_sample_stages()

    wb = Workbook()
    ws = wb.active
    ws.title = "Target Selection"

    headers = [
        "Sample", "Notes from first experiment",
        "Before-occlusion stage", "Before-occlusion image",
        "After-reappearance stage", "After-reappearance image",
        "Target description", "Box number before occlusion", "Box number after reappearance",
        "Same physical object? Yes/No", "Hidden, not left frame? Yes/No",
        "Include in study? Yes/No", "Rejection reason (if not included)",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    print(f"Rendering numbered before/after images at confidence >= {DISPLAY_CONF_THRESHOLD} "
          f"(decluttering only -- raw detections are untouched in all_detections.csv).\n")

    for sample, stages in stages_by_sample.items():
        before_candidates = [s for s in stages if s < 3]
        after_candidates = [s for s in stages if s > 3]
        if not before_candidates or not after_candidates:
            print(f"{sample}: skipped -- no valid before/after pair around full occlusion "
                  f"(stages present: {stages})")
            continue
        stage_before = max(before_candidates)
        stage_after = min(after_candidates)

        before_src = stage_image_path(sample, stage_before)
        after_src = stage_image_path(sample, stage_after)

        sample_out_dir = SELECTION_IMAGES_ROOT / sample
        before_out = sample_out_dir / f"before_stage{stage_before}_{STAGE_NAMES[stage_before]}.jpg"
        after_out = sample_out_dir / f"after_stage{stage_after}_{STAGE_NAMES[stage_after]}.jpg"

        before_dets = detections_by_key.get((sample, stage_before), [])
        after_dets = detections_by_key.get((sample, stage_after), [])

        before_shown = draw_numbered_boxes(before_src, before_dets, before_out)
        after_shown = draw_numbered_boxes(after_src, after_dets, after_out)

        print(f"{sample}: before={STAGE_NAMES[stage_before]} ({len(before_shown)} numbered boxes), "
              f"after={STAGE_NAMES[stage_after]} ({len(after_shown)} numbered boxes)")

        ws.append([
            sample,
            KNOWN_ISSUE_NOTES.get(sample, ""),
            STAGE_NAMES[stage_before], str(before_out.relative_to(PROJECT_ROOT)).replace("\\", "/"),
            STAGE_NAMES[stage_after], str(after_out.relative_to(PROJECT_ROOT)).replace("\\", "/"),
            "", "", "", "", "", "", "",
        ])

    last_row = ws.max_row

    def add_dropdown(col_letter: str, options: list[str]):
        dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"',
                              allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{last_row}")

    add_dropdown("J", ["Yes", "No"])   # Same physical object?
    add_dropdown("K", ["Yes", "No"])   # Hidden, not left frame?
    add_dropdown("L", ["Yes", "No"])   # Include in study?

    widths = [12, 34, 16, 42, 18, 42, 26, 14, 14, 16, 18, 16, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"

    WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(WORKBOOK_PATH)
    print(f"\nWrote {WORKBOOK_PATH} with {last_row - 1} candidate sample rows.")


if __name__ == "__main__":
    main()
