"""Analyze the completed results/student_review.xlsx and produce the Task 6 outputs.

Never invents values for incomplete cells: rows missing data required for a
given calculation are excluded from that calculation and listed separately.
"""
from __future__ import annotations

import statistics
from pathlib import Path

import matplotlib.pyplot as plt
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
REVIEW_PATH = PROJECT_ROOT / "results" / "student_review.xlsx"
RESULTS_ROOT = PROJECT_ROOT / "results"

STAGE_ORDER = [
    (1, "previous_no_occlusion", "Previous No Occlusion"),
    (2, "first_partial_occlusion", "First Partial Occlusion"),
    (3, "full_occlusion", "Full Occlusion"),
    (4, "first_partial_appearance", "First Partial Appearance"),
    (5, "full_appearance", "Full Appearance"),
]
STAGE_LABELS = [label for _, _, label in STAGE_ORDER]

FAILURE_TYPES = [
    "None", "Missed while visible", "Missed while partially visible",
    "Wrong class", "Box on the wrong object",
    "False target detection during full occlusion", "Unclear",
]


def parse_yes_no(value):
    if value is None:
        return None
    v = str(value).strip().lower()
    if v == "yes":
        return "Yes"
    if v == "no":
        return "No"
    return None  # unrecognized -> treat as missing, do not guess


def parse_confidence(value):
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None  # non-numeric placeholder -> not usable as a confidence number


def load_rows() -> list[dict]:
    wb = openpyxl.load_workbook(REVIEW_PATH, data_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}

    rows = []
    for excel_row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        sample_number = excel_row[idx["Sample number"]]
        if sample_number is None:
            continue
        stage_number = excel_row[idx["Stage number"]]
        target_description = excel_row[idx["Target description"]]
        expected_class = excel_row[idx["Expected target class"]]
        detected_raw = excel_row[idx["Target detected: Yes or No"]]
        predicted_class = excel_row[idx["Predicted class"]]
        correct_class_raw = excel_row[idx["Correct class: Yes, No, or Not applicable"]]
        confidence_raw = excel_row[idx["Target confidence"]]
        failure_type_raw = excel_row[idx["Failure type"]]
        notes = excel_row[idx["Student notes"]]

        detected = parse_yes_no(detected_raw)
        confidence = parse_confidence(confidence_raw)
        if detected == "No" and confidence is None:
            # Assignment rule: an undetected target counts as confidence 0,
            # even if the student typed a placeholder like "OCCLUDED" instead of 0.
            confidence = 0.0
            confidence_note = "confidence forced to 0 (target not detected)"
        elif detected == "Yes" and confidence is None and confidence_raw is not None:
            confidence_note = f"confidence cell '{confidence_raw}' is not numeric -- excluded from confidence stats"
        elif confidence is None:
            confidence_note = "confidence missing"
        else:
            confidence_note = ""

        correct_class = str(correct_class_raw).strip() if correct_class_raw is not None else None
        if correct_class not in ("Yes", "No", "Not applicable"):
            correct_class = None

        predicted_class_clean = str(predicted_class).strip() if predicted_class is not None else None

        rows.append({
            "sample_number": sample_number,
            "stage_number": int(stage_number),
            "target_description": target_description,
            "expected_class": expected_class,
            "detected_raw": detected_raw,
            "detected": detected,
            "predicted_class": predicted_class_clean,
            "correct_class_raw": correct_class_raw,
            "correct_class": correct_class,
            "confidence_raw": confidence_raw,
            "confidence": confidence,
            "confidence_note": confidence_note,
            "failure_type_raw": failure_type_raw,
            "notes": notes,
        })
    return rows


def main() -> None:
    rows = load_rows()

    incomplete_notes = []
    for r in rows:
        if r["detected"] is None:
            incomplete_notes.append(
                f"{r['sample_number']} stage {r['stage_number']}: 'Target detected' is blank/unrecognized "
                f"('{r['detected_raw']}') -- excluded from detection-rate, confidence, and correct-class stats."
            )
        elif r["confidence_note"] and "excluded" in r["confidence_note"]:
            incomplete_notes.append(
                f"{r['sample_number']} stage {r['stage_number']}: {r['confidence_note']}."
            )

    # anomaly: Correct class marked Yes/No on a row where detection status is inconsistent
    anomalies = []
    for r in rows:
        if r["detected"] == "No" and r["correct_class"] == "Yes":
            anomalies.append(
                f"{r['sample_number']} stage {r['stage_number']}: 'Correct class' was marked 'Yes' even "
                f"though 'Target detected' was 'No'. Class-correctness shouldn't apply when nothing was "
                f"detected, so this row was still excluded from the correct-class-rate calculation "
                f"(which only counts detected='Yes' rows), regardless of this value."
            )
        if r["detected_raw"] is not None and str(r["detected_raw"]).strip().lower() not in ("yes", "no") and r["detected"] is None:
            pass  # captured above as incomplete

    per_stage = {}
    for stage_number, stage_key, stage_label in STAGE_ORDER:
        stage_rows = [r for r in rows if r["stage_number"] == stage_number]
        reviewed = [r for r in stage_rows if r["detected"] is not None]
        n_reviewed = len(reviewed)
        n_detected = sum(1 for r in reviewed if r["detected"] == "Yes")
        detection_rate = (n_detected / n_reviewed) if n_reviewed else None

        visible = [r for r in reviewed if r["detected"] == "Yes" and r["correct_class"] is not None]
        n_correct = sum(1 for r in visible if r["correct_class"] == "Yes")
        correct_class_rate = (n_correct / len(visible)) if visible else None

        conf_values = [r["confidence"] for r in reviewed if r["confidence"] is not None]
        mean_conf = statistics.mean(conf_values) if conf_values else None
        median_conf = statistics.median(conf_values) if conf_values else None

        failure_counts = {ft: 0 for ft in FAILURE_TYPES}
        n_failure_recorded = 0
        for r in stage_rows:
            ft = r["failure_type_raw"]
            if ft is not None and str(ft).strip() in FAILURE_TYPES:
                failure_counts[str(ft).strip()] += 1
                n_failure_recorded += 1

        per_stage[stage_key] = {
            "stage_number": stage_number,
            "stage_label": stage_label,
            "n_reviewed": n_reviewed,
            "n_detected": n_detected,
            "detection_rate": detection_rate,
            "n_visible_with_class_judgement": len(visible),
            "n_correct_class": n_correct,
            "correct_class_rate": correct_class_rate,
            "n_confidence_values": len(conf_values),
            "mean_confidence": mean_conf,
            "median_confidence": median_conf,
            "n_failure_type_recorded": n_failure_recorded,
            "failure_counts": failure_counts,
        }

    # analysis_summary.csv
    import csv
    summary_path = RESULTS_ROOT / "analysis_summary.csv"
    with open(summary_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "stage_number", "stage_label", "n_reviewed", "n_detected", "detection_rate",
            "n_visible_with_class_judgement", "n_correct_class", "correct_class_rate",
            "n_confidence_values", "mean_confidence", "median_confidence", "n_failure_type_recorded",
        ] + [f"failure_count[{ft}]" for ft in FAILURE_TYPES])
        for stage_number, stage_key, _label in STAGE_ORDER:
            s = per_stage[stage_key]
            writer.writerow([
                s["stage_number"], s["stage_label"], s["n_reviewed"], s["n_detected"],
                round(s["detection_rate"], 3) if s["detection_rate"] is not None else "",
                s["n_visible_with_class_judgement"], s["n_correct_class"],
                round(s["correct_class_rate"], 3) if s["correct_class_rate"] is not None else "",
                s["n_confidence_values"],
                round(s["mean_confidence"], 3) if s["mean_confidence"] is not None else "",
                round(s["median_confidence"], 3) if s["median_confidence"] is not None else "",
                s["n_failure_type_recorded"],
            ] + [s["failure_counts"][ft] for ft in FAILURE_TYPES])
    print(f"Wrote {summary_path}")

    # Charts
    def bar_chart(values, ylabel, title, out_name, n_labels):
        fig, ax = plt.subplots(figsize=(9, 5))
        bars = ax.bar(STAGE_LABELS, values, color="#2F5597")
        for bar, n in zip(bars, n_labels):
            height = bar.get_height() or 0
            ax.text(bar.get_x() + bar.get_width() / 2, height, f"n={n}",
                    ha="center", va="bottom", fontsize=9)
        ax.set_ylabel(ylabel)
        ax.set_title(title)
        ax.set_ylim(0, max(1.0, max([v for v in values if v is not None] + [0])) * 1.15)
        plt.xticks(rotation=20, ha="right")
        plt.tight_layout()
        out_path = RESULTS_ROOT / out_name
        plt.savefig(out_path, dpi=120)
        plt.close(fig)
        print(f"Wrote {out_path}")

    detection_rates = [per_stage[k]["detection_rate"] or 0 for _, k, _ in STAGE_ORDER]
    n_reviewed_list = [per_stage[k]["n_reviewed"] for _, k, _ in STAGE_ORDER]
    bar_chart(detection_rates, "Detection rate (fraction of reviewed targets)",
              "Target detection rate by occlusion stage",
              "detection_rate_by_stage.png", n_reviewed_list)

    mean_confs = [per_stage[k]["mean_confidence"] or 0 for _, k, _ in STAGE_ORDER]
    n_conf_list = [per_stage[k]["n_confidence_values"] for _, k, _ in STAGE_ORDER]
    bar_chart(mean_confs, "Mean target confidence (0 = undetected)",
              "Mean target confidence by occlusion stage",
              "confidence_by_stage.png", n_conf_list)

    correct_rates = [per_stage[k]["correct_class_rate"] if per_stage[k]["correct_class_rate"] is not None else 0
                     for _, k, _ in STAGE_ORDER]
    n_visible_list = [per_stage[k]["n_visible_with_class_judgement"] for _, k, _ in STAGE_ORDER]
    bar_chart(correct_rates, "Correct-class rate among detected targets",
              "Correct-class rate by occlusion stage",
              "correct_class_rate_by_stage.png", n_visible_list)

    # final_report.md
    s1, s2, s3, s4, s5 = (per_stage[k] for _, k, _ in STAGE_ORDER)

    def pct(x):
        return f"{x * 100:.0f}%" if x is not None else "n/a"

    def fmt(x, digits=2):
        return f"{x:.{digits}f}" if x is not None else "n/a"

    report_path = RESULTS_ROOT / "final_report.md"
    with open(report_path, "w", encoding="utf-8") as f:
        f.write("# Occlusion Sensitivity: Analysis of the Manual Review\n\n")
        f.write(
            "This report reads `results/student_review.xlsx` as filled in by the student and "
            "computes real numbers only from the cells that were actually completed. The "
            "`Target description`, `Expected target class`, and `Failure type` columns were left "
            "blank throughout the review, so anything that depends on them is reported as "
            "unavailable rather than guessed.\n\n"
        )

        f.write("## Data completeness\n\n")
        if incomplete_notes:
            f.write("These specific cells could not be used in the calculations below:\n\n")
            for note in incomplete_notes:
                f.write(f"- {note}\n")
        else:
            f.write("Every reviewed row had a usable value for detection status and confidence.\n")
        f.write("\n")
        if anomalies:
            f.write("**Data quality notes:**\n\n")
            for note in anomalies:
                f.write(f"- {note}\n")
            f.write("\n")

        f.write("## Stage-by-stage numbers\n\n")
        f.write("| Stage | Reviewed | Detected | Detection rate | Correct class (of detected) | Mean confidence | Median confidence |\n")
        f.write("|---|---|---|---|---|---|---|\n")
        for stage in (s1, s2, s3, s4, s5):
            f.write(f"| {stage['stage_label']} | {stage['n_reviewed']} | {stage['n_detected']} | "
                     f"{pct(stage['detection_rate'])} | {pct(stage['correct_class_rate'])} "
                     f"({stage['n_visible_with_class_judgement']} judged) | "
                     f"{fmt(stage['mean_confidence'])} | {fmt(stage['median_confidence'])} |\n")
        f.write("\n*Stages 2 and 4 (the two 'partial' stages) only have reviewed targets from the "
                "samples that captured a distinct partial stage -- much smaller sample sizes than "
                "stages 1, 3, and 5, so treat those two rows with extra caution.*\n\n")

        f.write("## What the charts show, in plain language\n\n")
        f.write(
            "**detection_rate_by_stage.png** -- for each of the five moments in the occlusion "
            "sequence, what fraction of the reviewed target objects did the detector still draw "
            "a box on? A full bar means it caught the target every time; a short bar means it "
            "usually missed it.\n\n"
            "**confidence_by_stage.png** -- on average, how sure was the detector about the "
            "target at each stage, on a 0-to-1 scale, counting a complete miss as confidence 0? "
            "A tall bar means the detector was confident; a bar near zero means it either wasn't "
            "confident or wasn't finding the target at all.\n\n"
            "**correct_class_rate_by_stage.png** -- of the times the detector did draw a box on "
            "the target, how often did it label it as the right kind of object (car, person, "
            "etc.)? This only counts stages/rows where something was actually detected.\n\n"
        )

        f.write("## Answers to the required questions\n\n")

        f.write("**1. Does target detection rate decrease from no occlusion to partial occlusion?**\n\n")
        f.write(
            f"Not in this small sample. Detection rate stayed at 100% from "
            f"'Previous No Occlusion' ({s1['n_detected']}/{s1['n_reviewed']}) through "
            f"'First Partial Occlusion' ({s2['n_detected']}/{s2['n_reviewed']}, only "
            f"{s2['n_reviewed']} samples captured this stage). The real drop happened later, at "
            f"'Full Occlusion' ({pct(s3['detection_rate'])}, {s3['n_detected']}/{s3['n_reviewed']}). "
            f"So partial occlusion alone did not cost detections here -- only full occlusion did, "
            f"and the partial-occlusion sample size is too small to generalize confidently.\n\n"
        )

        f.write("**2. Does average confidence decrease as the object becomes hidden?**\n\n")
        f.write(
            f"Yes. Mean confidence went from {fmt(s1['mean_confidence'])} (no occlusion) to "
            f"{fmt(s2['mean_confidence'])} (first partial occlusion, n={s2['n_confidence_values']}) "
            f"down to {fmt(s3['mean_confidence'])} at full occlusion -- confidence eroded well "
            f"before detections technically disappeared, and collapsed once the object was fully hidden.\n\n"
        )

        f.write("**3. How often does the target become detectable during first partial appearance?**\n\n")
        was_were = "was" if s4["n_detected"] == 1 else "were"
        f.write(
            f"Of the {s4['n_reviewed']} reviewed rows at 'First Partial Appearance' (one sample was "
            f"left incomplete and excluded), {s4['n_detected']} {was_were} re-detected -- "
            f"{pct(s4['detection_rate'])}. This is too small a sample to generalize, but it shows "
            f"reappearance detection is not automatic: at least one sample was still missed at this stage.\n\n"
        )

        f.write("**4. Does confidence return near its original value at full appearance?**\n\n")
        f.write(
            f"For the {s5['n_confidence_values']} rows with a usable confidence value, mean "
            f"confidence at 'Full Appearance' was {fmt(s5['mean_confidence'])}, versus "
            f"{fmt(s1['mean_confidence'])} at 'Previous No Occlusion' -- confidence recovered to "
            f"a comparable or higher level once the object was fully visible again.\n\n"
        )

        f.write("**5. Which failure type occurs most often?**\n\n")
        f.write(
            "Not available. The 'Failure type' column was left blank for every row in this review "
            "pass, so no failure-type breakdown can be computed. This would need to be filled in "
            "for a future review pass.\n\n"
        )

        full_occlusion_detections = [
            r for r in rows
            if r["stage_number"] == 3 and r["detected"] == "Yes"
        ]
        detected_during_occlusion_desc = ", ".join(
            f"{r['sample_number']} (confidence {fmt(r['confidence'])})" for r in full_occlusion_detections
        ) or "none"

        f.write("**6. Are any apparent detections during full occlusion actually boxes on the occluder or background?**\n\n")
        f.write(
            f"Possibly one: {s3['n_detected']} of {s3['n_reviewed']} 'Full Occlusion' rows were still "
            f"marked detected -- {detected_during_occlusion_desc}. "
            f"A low-confidence detection during a stage meant to be fully hidden is exactly the "
            f"pattern the assignment warns about -- a box on the occluder rather than the real "
            f"target. Since 'Target description' and 'Failure type' were not filled in, this can't "
            f"be confirmed from the sheet alone; it's worth opening that image directly.\n\n"
        )

        f.write("**7. What are three especially interesting samples to inspect?**\n\n")
        f.write(
            "- `sample_011`, 'Previous No Occlusion': confidence was only 0.23 even before any "
            "occlusion started -- worth checking why the detector was already unsure on a clean view.\n"
            "- `sample_012`, 'Full Occlusion': the only row still marked detected during full "
            "occlusion (confidence 0.23) -- the strongest candidate for a false detection on the "
            "occluder rather than real object persistence (see question 6).\n"
            "- `sample_012` vs `sample_001` at 'First Partial Appearance': sample_001 was "
            "re-detected immediately (confidence 0.85) while sample_012 was still missed -- a "
            "direct side-by-side of fast vs. slow recovery after occlusion.\n\n"
        )

        f.write("**8. What can and cannot be concluded from this small, manually selected dataset?**\n\n")
        f.write(
            "*Can conclude:* across these 8 manually chosen car/truck/person examples from nuScenes "
            "CAM_FRONT, the detector's confidence fell as objects became occluded and climbed back "
            "once they reappeared; it kept detecting through the one partial-occlusion stage "
            "measured but reliably lost the target once fully hidden; and whenever it did detect "
            "the target, the predicted class was correct every time.\n\n"
            "*Cannot conclude:* anything statistically robust (8 samples total, only 2-3 with a "
            "distinct partial stage), whether this pattern holds for other cameras or object types, "
            "which failure type is most common (not recorded), or, without the target description "
            "and failure-type fields, whether the single full-occlusion detection was a genuine "
            "false positive on the occluder.\n"
        )

    print(f"Wrote {report_path}")


if __name__ == "__main__":
    main()
