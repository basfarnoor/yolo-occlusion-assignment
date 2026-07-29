"""Fill Claude's proposed target picks into the Target Selection sheet.
These are automated best-guesses (class + position matching across frames) --
clearly a proposal for the student to visually confirm or override, not a
final decision. See assignments/02_last_seen_memory/README.md Task 1.
"""
import openpyxl

WORKBOOK_PATH = r"C:\Users\HP\OneDrive\Desktop\srsi\results\last_seen_memory\last_seen_experiment.xlsx"

# sample -> (description, box_before, box_after, same_object, hidden, include, reason_if_excluded)
PROPOSALS = {
    "sample_004": (
        "Light-colored car crossing the intersection ahead, center-right of frame",
        3, 1, "Yes", "Yes", "Yes", ""),
    "sample_011": (
        "Pedestrian standing near the middle of the crosswalk",
        3, 4, "Yes", "Yes", "Yes", ""),
    "sample_006": (
        "Car in the middle of the road, center of frame",
        2, 1, "Yes", "Yes", "Yes", ""),
    "sample_005": (
        "Truck on the right side of the intersection",
        3, 3, "Yes", "Yes", "Yes", ""),
    "sample_003": (
        "Car near the center of the intersection",
        1, 4, "Yes", "Yes", "Yes", ""),
    "sample_001": (
        "Car near center-left of the street (weaker match, proposed as backup only)",
        3, 2, "Yes", "Yes", "No",
        "Position match is looser (~120px drift) than the 5 samples already proposed -- "
        "kept as a possible backup, not one of the primary 5."),
    "sample_012": (
        "Pedestrian near the crosswalk on the right (low detection confidence, 0.41)",
        4, 2, "", "",  "No",
        "Best auto-matched candidate is a low-confidence pedestrian with ~125px drift. "
        "The large silver SUV in this scene (box 1 in both selection images) never actually "
        "disappears at full occlusion, so it doesn't qualify. Worth a manual look given this "
        "sample's earlier flagged false-detection issue, but no clean auto-match found."),
    "sample_007": (
        "Truck near the intersection (auto-match uncertain)",
        3, 4, "", "", "No",
        "Best auto-matched candidate drifted ~475px between before/after -- too large to trust "
        "as the same physical object without closer visual inspection."),
}

wb = openpyxl.load_workbook(WORKBOOK_PATH)
ws = wb["Target Selection"]
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
col = {name: i + 1 for i, name in enumerate(headers)}

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    sample = row[0].value
    if sample not in PROPOSALS:
        continue
    desc, box_before, box_after, same_obj, hidden, include, reason = PROPOSALS[sample]
    row[col["Target description"] - 1].value = desc
    row[col["Box number before occlusion"] - 1].value = box_before
    row[col["Box number after reappearance"] - 1].value = box_after
    row[col["Same physical object? Yes/No"] - 1].value = same_obj
    row[col["Hidden, not left frame? Yes/No"] - 1].value = hidden
    row[col["Include in study? Yes/No"] - 1].value = include
    row[col["Rejection reason (if not included)"] - 1].value = reason

wb.save(WORKBOOK_PATH)
print(f"Wrote proposed picks for {len(PROPOSALS)} samples into {WORKBOOK_PATH}")
