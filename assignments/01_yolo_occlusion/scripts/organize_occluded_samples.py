"""Copy nuScenes images referenced in occlusion_samples.xlsx into occluded_samples/.

Safe to rerun: never overwrites, never deletes, never moves/edits source files.
See assignments/01_yolo_occlusion/README.md Task 3 for the full specification.
"""
from __future__ import annotations

import csv
import hashlib
import os
import shutil
import sys
from dataclasses import dataclass, field

import openpyxl

ASSIGNMENT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REPO_ROOT = os.path.dirname(os.path.dirname(ASSIGNMENT_ROOT))
DATA_ROOT = os.path.join(REPO_ROOT, "data")
WORKBOOK_PATH = os.path.join(REPO_ROOT, "occlusion_samples.xlsx")
OUTPUT_ROOT = os.path.join(ASSIGNMENT_ROOT, "occluded_samples")
IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png")

STAGES = [
    (1, "previous_no_occlusion", ["Previous No Occlusion"]),
    (2, "first_partial_occlusion", ["First Partial Occlusion"]),
    (3, "full_occlusion", ["Full Occlusion"]),
    (4, "first_partial_appearance", ["First Partial Apparence", "First Partial Appearence", "First Partial Appearance"]),
    (5, "full_appearance", ["Full Appearence", "Full Appearance"]),
]

DASH_PLACEHOLDERS = {"-", "--", "—", ""}


@dataclass
class ManifestRow:
    sample_number: str
    excel_row: int
    stage_number: int
    stage_name: str
    excel_value: str
    source_path: str
    destination_path: str
    status: str
    notes: str = ""


def normalize_header(name: str) -> str:
    return " ".join(str(name).strip().split()).lower()


def find_column_map(header_row) -> dict[str, int]:
    """Map each of our 6 logical columns to its 1-based column index."""
    wanted = {
        "sample no.": "sample_number",
    }
    for _, key, aliases in STAGES:
        for alias in aliases:
            wanted[normalize_header(alias)] = key

    col_map = {}
    for idx, cell in enumerate(header_row, start=1):
        if cell.value is None:
            continue
        norm = normalize_header(cell.value)
        if norm in wanted:
            col_map[wanted[norm]] = idx
    return col_map


def build_filename_index(root: str) -> dict[str, list[str]]:
    """stem (no extension) -> list of relative paths under DATA_ROOT with that stem."""
    index: dict[str, list[str]] = {}
    for dirpath, _dirnames, filenames in os.walk(root):
        for fname in filenames:
            stem, ext = os.path.splitext(fname)
            if ext.lower() not in IMAGE_EXTENSIONS:
                continue
            rel = os.path.relpath(os.path.join(dirpath, fname), root)
            index.setdefault(stem, []).append(rel)
    return index


def build_lowercase_index(index: dict[str, list[str]]) -> dict[str, list[str]]:
    lower_index: dict[str, list[str]] = {}
    for stem, paths in index.items():
        lower_index.setdefault(stem.lower(), []).extend(paths)
    return lower_index


def resolve_image(raw_value, index: dict[str, list[str]], lower_index: dict[str, list[str]]):
    """Return (status, source_rel_path_or_None, notes)."""
    if raw_value is None:
        return "empty_excel_cell", None, ""

    value = str(raw_value).strip()
    if value == "" or value in DASH_PLACEHOLDERS:
        return "missing", None, f"Cell contains placeholder {raw_value!r}, not a real filename."

    # 1) exact relative path provided
    candidate = os.path.join(DATA_ROOT, value)
    if os.path.isfile(candidate):
        return "copied", value.replace("\\", "/"), ""

    # 2) exact filename (with extension already) search
    base = os.path.basename(value)
    stem, ext = os.path.splitext(base)
    if ext.lower() in IMAGE_EXTENSIONS and stem in index:
        matches = index[stem]
        if len(matches) == 1:
            return "copied", matches[0].replace("\\", "/"), ""
        return "ambiguous", None, "Multiple files with this exact name: " + ", ".join(matches)

    # 3) exact stem + supported extension
    if base in index:
        matches = index[base]
        if len(matches) == 1:
            return "copied", matches[0].replace("\\", "/"), ""
        return "ambiguous", None, "Multiple files with this exact stem: " + ", ".join(matches)

    # 4) case-insensitive fallback
    lower_key = base.lower()
    if lower_key in lower_index:
        matches = lower_index[lower_key]
        if len(matches) == 1:
            return "copied", matches[0].replace("\\", "/"), "Matched case-insensitively."
        return "ambiguous", None, "Multiple case-insensitive matches: " + ", ".join(matches)

    return "missing", None, f"No file named '{value}' (with a supported extension) found anywhere under data/."


def sha256_of(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def copy_verified(src_abs: str, dst_abs: str) -> tuple[bool, str]:
    os.makedirs(os.path.dirname(dst_abs), exist_ok=True)
    if os.path.exists(dst_abs):
        if os.path.getsize(dst_abs) == os.path.getsize(src_abs) and sha256_of(dst_abs) == sha256_of(src_abs):
            return True, "already_present"
        return False, "conflict"
    shutil.copy2(src_abs, dst_abs)
    if os.path.getsize(dst_abs) != os.path.getsize(src_abs) or sha256_of(dst_abs) != sha256_of(src_abs):
        return False, "conflict"
    return True, "copied"


def main() -> None:
    if not os.path.isfile(WORKBOOK_PATH):
        print(f"ERROR: workbook not found at {WORKBOOK_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row))
    header_row, data_rows = rows[0], rows[1:]

    col_map = find_column_map(header_row)
    required = {"sample_number"} | {key for _, key, _ in STAGES}
    missing_cols = required - set(col_map)
    if missing_cols:
        print("ERROR: could not find expected columns:", missing_cols)
        sys.exit(1)

    print("Scanning data/ for image files (this only reads filenames, nothing is copied yet)...")
    index = build_filename_index(DATA_ROOT)
    lower_index = build_lowercase_index(index)
    print(f"Indexed {sum(len(v) for v in index.values())} image files under data/.")

    os.makedirs(OUTPUT_ROOT, exist_ok=True)

    manifest: list[ManifestRow] = []
    seen_sample_numbers: dict[str, list[int]] = {}

    for seq, row in enumerate(data_rows, start=1):
        excel_row_num = row[0].row
        sample_no_cell = row[col_map["sample_number"] - 1].value
        # normalize e.g. 4.0 -> "4"
        try:
            sample_label = str(int(float(sample_no_cell)))
        except (TypeError, ValueError):
            sample_label = str(sample_no_cell).strip() if sample_no_cell is not None else "(blank)"

        seen_sample_numbers.setdefault(sample_label, []).append(excel_row_num)

        folder_name = f"sample_{seq:03d}"
        dest_dir = os.path.join(OUTPUT_ROOT, folder_name)
        os.makedirs(dest_dir, exist_ok=True)

        for stage_number, stage_key, _aliases in STAGES:
            col_idx = col_map[stage_key]
            raw_value = row[col_idx - 1].value
            stage_name = stage_key

            status, rel_source, notes = resolve_image(raw_value, index, lower_index)
            dest_path = ""
            source_path_display = rel_source or ""

            if status == "copied":
                src_abs = os.path.join(DATA_ROOT, rel_source)
                _stem, ext = os.path.splitext(rel_source)
                dest_filename = f"{stage_number}_{stage_name}{ext.lower()}"
                dest_abs = os.path.join(dest_dir, dest_filename)
                dest_path = os.path.join("occluded_samples", folder_name, dest_filename).replace("\\", "/")
                ok, copy_status = copy_verified(src_abs, dest_abs)
                if copy_status == "already_present":
                    status = "already_present"
                elif copy_status == "conflict":
                    status = "conflict"
                    notes = (notes + " " if notes else "") + "Destination existed with different content/size; left untouched."
                else:
                    status = "copied"
                source_path_display = rel_source

            manifest.append(ManifestRow(
                sample_number=sample_label,
                excel_row=excel_row_num,
                stage_number=stage_number,
                stage_name=stage_name,
                excel_value="" if raw_value is None else str(raw_value),
                source_path=source_path_display,
                destination_path=dest_path,
                status=status,
                notes=notes,
            ))

    # Write manifest.csv
    manifest_path = os.path.join(OUTPUT_ROOT, "manifest.csv")
    with open(manifest_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["sample_number", "excel_row", "stage_number", "stage_name",
                          "excel_value", "source_path", "destination_path", "status", "notes"])
        for m in manifest:
            writer.writerow([m.sample_number, m.excel_row, m.stage_number, m.stage_name,
                              m.excel_value, m.source_path, m.destination_path, m.status, m.notes])

    # Build report
    from collections import Counter
    status_counts = Counter(m.status for m in manifest)
    n_rows = len(data_rows)
    n_expected_images = len(manifest)
    duplicates = {k: v for k, v in seen_sample_numbers.items() if len(v) > 1}

    attention_rows = [m for m in manifest if m.status in ("missing", "ambiguous", "conflict")]

    report_path = os.path.join(OUTPUT_ROOT, "organization_report.md")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write("# Organization Report\n\n")
        f.write(f"- Workbook sample rows: **{n_rows}**\n")
        f.write(f"- Expected images (rows x 5 stages): **{n_expected_images}**\n")
        f.write(f"- Copied: **{status_counts.get('copied', 0)}**\n")
        f.write(f"- Already present (rerun-safe skip): **{status_counts.get('already_present', 0)}**\n")
        f.write(f"- Missing: **{status_counts.get('missing', 0)}**\n")
        f.write(f"- Ambiguous: **{status_counts.get('ambiguous', 0)}**\n")
        f.write(f"- Conflicting: **{status_counts.get('conflict', 0)}**\n")
        f.write(f"- Empty Excel cells: **{status_counts.get('empty_excel_cell', 0)}**\n\n")

        f.write("## Duplicate sample numbers in the workbook\n\n")
        if duplicates:
            for label, excel_rows in duplicates.items():
                f.write(f"- Sample number `{label}` appears in Excel rows {excel_rows} "
                        f"({len(excel_rows)} occurrences). Each was organized as its own separate "
                        f"folder (by row order) since the underlying filenames differ.\n")
        else:
            f.write("None found.\n")
        f.write("\n")

        f.write("## Rows requiring human attention\n\n")
        if attention_rows:
            f.write("| Sample no. | Excel row | Stage | Excel value | Status | Notes |\n")
            f.write("|---|---|---|---|---|---|\n")
            for m in attention_rows:
                f.write(f"| {m.sample_number} | {m.excel_row} | {m.stage_name} | "
                        f"{m.excel_value} | {m.status} | {m.notes} |\n")
        else:
            f.write("None. Every non-empty cell resolved to exactly one file.\n")
        f.write("\n")

        f.write("## Folder mapping\n\n")
        f.write("Folders are numbered by the workbook's row order (`sample_001`, `sample_002`, ...), "
                "not by the Excel `Sample no.` value, because that value repeats for rows 4 and 8. "
                "The manifest's `sample_number` and `excel_row` columns record the original label "
                "and exact row for traceability.\n\n")

        f.write("## Notes for the student\n\n")
        f.write("- Sample 1's 'Previous No Occlusion' and 'First Partial Occlusion' cells contained "
                "the identical filename in the workbook; both stages were copied as that same source "
                "image, so those two files in `sample_001/` will look identical. Confirm this was intentional.\n")
        f.write("- Rows with only a dash placeholder (samples 2, 5, 6, 7) and the fully blank rows "
                "(samples 9, 10) produced folders containing zero or very few images; see the manifest.\n")

    print(f"Done. Wrote {manifest_path} and {report_path}.")
    print(f"Status counts: {dict(status_counts)}")


if __name__ == "__main__":
    main()
